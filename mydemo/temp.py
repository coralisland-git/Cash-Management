from django.shortcuts import render_to_response, redirect
from django.http import HttpResponse
from django.template import Template, Context
from django.template import RequestContext
from django.contrib.auth.models import *
from mydemo.models import *
from functools import wraps
from django.shortcuts import render
import datetime
from django.http import HttpResponse
from myProject import settings
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl import Workbook 
import datetime 
from openpyxl.styles import colors
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
import pdb

# Create your views here.

def index(request):
    return HttpResponse("Welcome visit our page")

def login_required():
    def login_decorator(function):
        @wraps(function)
        def wrapped_function(request):

            # if a user is not authorized, redirect to login page
            if 'user' not in request.session or request.session['user'] is None:
                return redirect("/")
            # otherwise, go on the request
            else:
                return function(request)

        return wrapped_function

    return login_decorator


# login view
def login(request):
    error = 'none'
    request.session['user'] = None

    if 'username' in request.POST:

        # get username and password from request.
        username = request.POST['username']
        password = request.POST['password']
        # check whether the user is in database or not
        if username == settings.ADMIN_NAME and password == settings.ADMIN_PASSWORD:
            request.session['user'] = {
                # "id": user[0].id,
                "username": settings.ADMIN_NAME, #user[0].email,
                "password": settings.ADMIN_PASSWORD, #user[0].name.split(" ")[0],
                "role": "admin"
            }

            return redirect("/home")
        else :
            error = 'block'

    return render(request, 'login.html', locals())

    # return render_to_response('login.html', {'error':error}, context_instance=RequestContext(request))

def logout(request):
    request.session['user'] = None
    return redirect("/")

def main(request):
    # clients = Client.objects.all().order_by("-created_on")
    return render(request, 'black.html')
    # return render_to_response('blank.html', locals(), context_instance=RequestContext(request))


def home(request):

    global invoice_arr

    invoice_arr = []

    global timecard_arr

    global reconkey_arr

    global payment_arr

    return render(request, 'home.html')


def invoice_board(request):

    if request.method == 'POST' and request.FILES:        

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('uploaded_files/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading Invoice Data ...')

        input_invoice_data_sheet = workbook[workbook.sheetnames[0]]

        input_invoice_data_col_max_count = 0

        invoice_row = 1

        invoice_header = [
            'company', 
            'branch_id', 
            'employee_short_name', 
            'cust_alph', 
            'customer_name', 
            'invoice_number', 
            'straight_time_hours', 
            'hourly_bill_rate', 
            'overtime_hours', 
            'sp_bill', 
            'total_bill_amount', 
            'invoice_date', 
            'week_ending_date', 
            'invoice_reference', 
            'workers_comp_code', 
            'job_description', 
            'customer_po', 
            'b_w_location_code', 
            'address_1', 
            'spcl_paybill_de_code', 
            'special_paybill_qty', 
            'special_bill_rate'
         ]

        for invoice in input_invoice_data_sheet.rows:

            try:

                data = {}

                for invoice_col in range(0, len(invoice)) :

                    data[invoice_header[invoice_col]] = invoice[invoice_col].value


                if invoice_row > 2:

                    items = data['invoice_reference'].split('@')

                    invoice_arr.append(data)

                invoice_row += 1

            except:

                pass

        # pdb.set_trace()

        print('~~~~~~~~~~~', invoice_arr)
 
        # request.session['invoice_arr'] = json.dumps(invoice_arr) 

        # pdb.set_trace()

        # data_table = Template("""
        #     <div class="portlet light ">
        #         <div class="portlet-title">
        #             <div class="caption font-green">
        #                 <i class="icon-settings font-green"></i>
        #                 <span class="caption-subject bold uppercase">BATCH ID</span>
        #             </div>
        #             <div class="tools"> </div>
        #         </div>
        #         <div class="portlet-body">
        #             <table class="table table-striped table-bordered table-hover dt-responsive invoice_table" width="100%" id="sample_2">
        #                 <thead>
        #                     <tr>
        #                         <th></th>
        #                         <td className="all">invoice_number</td>
        #                         <td className="all">total_bill_amount</td>
        #                         <td className="min-tablet">invoice_reference</td>
        #                         <td className="min-tablet">invoice_date</td>
        #                         <td className="none">company</td>
        #                         <td className="none">branch_id</td>
        #                         <td className="none">employee_short_name</td>
        #                         <td className="none">cust_alph</td>
        #                         <td className="none">customer_name</td>
        #                         <td className="none">straight_time_hours</td>
        #                         <td className="none">hourly_bill_rate</td>
        #                         <td className="none">overtime_hours</td>
        #                         <td className="none">sp_bill</td>
        #                         <td className="none">week_ending_date</td>
        #                         <td className="none">workers_comp_code</td>
        #                         <td className="none">job_description</td>
        #                         <td className="none">customer_po</td>
        #                         <td className="none">b_w_location_code</td>
        #                         <td className="none">address_1</td>
        #                         <td className="none">spcl_paybill_de_code</td>
        #                         <td className="none">special_paybill_qty</td>
        #                         <td className="none">special_bill_rate</td>
        #                     </tr>
        #                 </thead>
        #                 <tbody>
        #                     {% for invoice in invoice_arr %}
        #                       <tr>
        #                         <th></th>
        #                         <td> {{ invoice.invoice_number }} </td> 
        #                         <td> {{ invoice.total_bill_amount }} </td> 
        #                         <td> {{ invoice.invoice_reference }} </td> 
        #                         <td> {{ invoice.invoice_date }} </td> 
        #                         <td> {{ invoice.company }} </td> 
        #                         <td> {{ invoice.branch_id }} </td> 
        #                         <td> {{ invoice.employee_short_name }} </td> 
        #                         <td> {{ invoice.cust_alph }} </td> 
        #                         <td> {{ invoice.customer_name }} </td> 
        #                         <td> {{ invoice.straight_time_hours }} </td> 
        #                         <td> {{ invoice.hourly_bill_rate }} </td> 
        #                         <td> {{ invoice.overtime_hours }} </td> 
        #                         <td> {{ invoice.sp_bill }} </td> 
        #                         <td> {{ invoice.week_ending_date }} </td> 
        #                         <td> {{ invoice.workers_comp_code }} </td> 
        #                         <td> {{ invoice.job_description }} </td> 
        #                         <td> {{ invoice.customer_po }} </td> 
        #                         <td> {{ invoice.b_w_location_code }} </td> 
        #                         <td> {{ invoice.address_1 }} </td> 
        #                         <td> {{ invoice.spcl_paybill_de_code }} </td> 
        #                         <td> {{ invoice.special_paybill_qty }} </td> 
        #                         <td> {{ invoice.special_bill_rate }} </td> 
        #                       </tr>
        #                     {% endfor %}
        #                 </tbody>
        #             </table>
        #         </div>
        #     </div>

        #     """)

        # html = data_table.render(Context({'invoice_arr': invoice_arr}))

        # return HttpResponse(html)

        return render(request, 'invoice/index.html', 
            {

                'invoice_arr' : invoice_arr,

            })

    else:
        
        return render(request, 'invoice/index.html',
            {

                # 'invoice_arr' : json.loads(request.session.get('invoice_arr', '[]')),

                'invoice_arr' : invoice_arr,

            })



def timecard_board(request):

    if request.POST :
        
        return render(request, 'timecard/index.html', {
            'uploaded' : 'uploaded'
            })

    return render(request, 'timecard/index.html', locals())


def reconkeys_board(request):

    return render(request, 'reconkeys/index.html', locals())


def payment_board(request):

    return render(request, 'payment/index.html', locals())


def signup(request):
    
    if request.POST:
        account = Account()
        account.firstname = request.POST["firstname"]
        account.lastname = request.POST["lastname"]
        account.username = request.POST["username"]
        account.email = request.POST["email"]
        account.phone = request.POST["phone"]
        account.address = request.POST["address"]
        account.password = request.POST["password"]
        request.session['user'] = {
	        # "id": user[0].id,
	        "firstname" : request.POST["firstname"],
	        "lastname" : request.POST["lastname"],
	        "username": request.POST["username"],
	        "email": request.POST["email"],
	        "phone" : request.POST["phone"],
	        "address" : request.POST["address"],
	        "password": request.POST["password"],
	        "role" : "client",
	    }
        account.save()
	   

        return redirect("/home")

    return render(request, 'signup.html')

    # return render_to_response('signup.html', locals(), context_instance=RequestContext(request))

def changePwd(request):
    if request.POST:
        if request.POST["opassword"] == request.session['user']['password']:
            if request.POST["password"] == request.POST["rpassword"]:
                account = Account.objects.filter(email=request.session['user']['email']).update(password=request.POST["password"])
                request.session['user']['password'] = request.POST['password']
                return redirect("/home") 


    return render(request, 'changePwd.html', locals())

	# return render_to_response("changePwd.html",locals(),context_instance=RequestContext(request))       

def forgot(request):
    if request.POST:
        if request.POST["password"] == request.POST["rpassword"]:
            account = Account.objects.filter(email=request.POST["email"]).update(password=request.POST["password"])
            return redirect("login") 

    return render(request, 'forgot.html')
    # return render_to_response("forgot.html",locals(),context_instance=RequestContext(request))           