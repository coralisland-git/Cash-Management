from django.shortcuts import render_to_response, redirect
from django.http import HttpResponse
from django.template import Template, Context
from django.template import RequestContext
from django.contrib.auth.models import *
from mydemo.models import *
from functools import wraps
from django.shortcuts import render
import datetime
from django.http import HttpResponse
from django.http import JsonResponse
from myProject import settings
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl import Workbook 
from openpyxl.styles import colors
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from django.db.models import Count
from django.forms.models import model_to_dict
import PyPDF2

import os
import pdb

# Create your views here.


def getvalue(in_str, first, last):

    ret = ''

    try:

        ret = in_str.split(first)[1].split(last)[0].strip()

    except:

        ret = ''

    return ret


def capitalize(item):

    try:

        item = item.replace('_', ' ').title()

        return item

    except:

        pass

def getPrice(item):

    try:

        item = float(item.strip().replace('$', ''))

        return item

    except :

        return 0

def getIndex(item, arr): 

    try:

        res = 0

        for ind in range(0, len(arr)):

            if arr[ind].strip().replace(':','') == item:

                res = ind

                break

        return res

    except :

        return 0


def index(request):
    return HttpResponse("Welcome visit our page")

def login_required():
    def login_decorator(function):
        @wraps(function)
        def wrapped_function(request):

            # if a user is not authorized, redirect to login page
            if 'user' not in request.session or request.session['user'] is None:
                return redirect("/")
            # otherwise, go on the request
            else:
                return function(request)

        return wrapped_function

    return login_decorator


# login view
def login(request):
    error = 'none'
    request.session['user'] = None

    if 'username' in request.POST:

        # get username and password from request.
        username = request.POST['username']
        password = request.POST['password']
        # check whether the user is in database or not
        if username == settings.ADMIN_NAME and password == settings.ADMIN_PASSWORD:
            request.session['user'] = {
                # "id": user[0].id,
                "username": settings.ADMIN_NAME, #user[0].email,
                "password": settings.ADMIN_PASSWORD, #user[0].name.split(" ")[0],
                "image": '',
                "firstname" : settings.ADMIN_FIRSTNAME ,
                'lastname' : settings.ADMIN_LASTNAME,
                "role": "admin",
                "logo" : settings.ADMIN_FIRSTNAME[0].upper() + settings.ADMIN_LASTNAME[0].upper()
            }

            return redirect("/home")
        else :
            error = 'block'

        user = Account.objects.filter(username=username, password=password) or Account.objects.filter(email=username, password=password)

        if len(user) > 0:
            request.session['user'] = {
                # "id": user[0].id,
                "firstname" : user[0].firstname,
                "lastname" : user[0].lastname,
                "username": user[0].username,
                "email": user[0].email,
                "password": user[0].password,
                "image" : user[0].image,
                "role" : "user",
                "logo" : user[0].firstname[0].upper() + user[0].lastname[0].upper()
            }

            return redirect("/home")

        else:

            error = 'block'

    return render(request, 'login.html', locals())

    # return render_to_response('login.html', {'error':error}, context_instance=RequestContext(request))

def logout(request):

    for key in list(request.session.keys()):
        
        del request.session[key]

    return redirect("/")


def clear_data(request):

    if 'refresh' in request.POST and request.POST['refresh'] == 'invoice_raw_data':

        request.session['invoice_arr'] = '[]'

        request.session['invoice_batch_no'] = ''

        request.session['invoice_view'] = 'detail'

        return redirect('/invoice_board')

    if 'refresh' in request.POST and request.POST['refresh'] == 'invoice_simple_data':

        request.session['input_invoice_arr'] = '[]'

        return redirect('/invoice_simple_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'timecard_kt_data':

        request.session['timecard_kt_arr'] = '[]'

        request.session['kt_tc_batch_no'] = ''

        request.session['kt_tc_view'] = 'detail'

        return redirect('/timecard_board')
        

    if 'refresh' in request.POST and request.POST['refresh'] == 'timecard_hb_data':

        request.session['timecard_hb_arr'] = '[]'

        request.session['hb_tc_batch_no'] = ''

        request.session['hb_tc_view'] = 'detail'

        return redirect('/timecard_hb_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'reconkeys_kt_data':

        request.session['reconkeys_arr'] = '[]'

        request.session['kt_rk_batch_no'] = ''

        request.session['kt_rk_view'] = 'detail'

        return redirect('/reconkeys_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'reconkeys_hb_data':

        request.session['reconkeys_hb_arr'] = '[]'

        request.session['hb_rk_batch_no'] = ''

        request.session['hb_rk_view'] = 'detail'

        return redirect('/reconkeys_hb_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'input_payment_data':

        request.session['input_payment_arr'] = '[]'

        request.session['payment_batch_no'] = ''

        request.session['payment_view'] = 'detail'

        return redirect('/payment_board')


def remove_data(request):

    if 'refresh' in request.POST and request.POST['refresh'] == 'invoice_raw_data':

        request.session['invoice_arr'] = '[]'

        Invoice.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['invoice_batch_no'] = ''

        request.session['invoice_view'] = 'batch'

        default_module(request)

        return redirect('/invoice_board')

    if 'refresh' in request.POST and request.POST['refresh'] == 'invoice_simple_data':

        request.session['input_invoice_arr'] = '[]'

        default_module(request)

        return redirect('/invoice_simple_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'timecard_kt_data':

        request.session['timecard_kt_arr'] = '[]'

        TimeCard_KT.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['kt_tc_batch_no'] = ''

        request.session['kt_tc_view'] = 'batch'

        default_module(request)

        return redirect('/timecard_board')
        

    if 'refresh' in request.POST and request.POST['refresh'] == 'timecard_hb_data':

        request.session['timecard_hb_arr'] = '[]'

        TimeCard_HB.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['hb_tc_batch_no'] = ''

        request.session['hb_tc_view'] = 'batch'

        default_module(request)

        return redirect('/timecard_hb_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'reconkeys_kt_data':

        request.session['reconkeys_arr'] = '[]'

        ReconKeys_KT.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['kt_rk_batch_no'] = ''

        request.session['kt_rk_view'] = 'batch'

        default_module(request)

        return redirect('/reconkeys_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'reconkeys_hb_data':

        request.session['reconkeys_hb_arr'] = '[]'

        ReconKeys_HB.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['hb_rk_batch_no'] = ''

        request.session['hb_rk_view'] = 'batch'

        default_module(request)

        return redirect('/reconkeys_hb_board')


    if 'refresh' in request.POST and request.POST['refresh'] == 'input_payment_data':

        request.session['input_payment_arr'] = '[]'

        Payment.objects.filter(batch_no=request.POST['batch_no']).delete()

        request.session['payment_batch_no'] = ''

        request.session['payment_view'] = 'batch'

        default_module(request)

        return redirect('/payment_board')


def search_batch_no(request):

    if 'invoice_batch_no' in request.POST:

        selected_invoices = Invoice.objects.filter(batch_no=str(request.POST['invoice_batch_no']))

        invoice_arr = []

        for invoice in selected_invoices:

            invoice_arr.append(model_to_dict(invoice))

        request.session['invoice_arr'] = json.dumps(invoice_arr)

        request.session['invoice_batch_no'] = str(request.POST['invoice_batch_no'])

        request.session['invoice_search_key'] = ''

        request.session['invoice_view'] = 'detail'
            
        return redirect('/invoice_board')


    if 'invoice_search_key' in request.POST:

        selected_invoices = []

        if str(request.POST['invoice_search_key']) == '' :

            selected_invoices = Invoice.objects.all()

        else:

            selected_invoices = Invoice.objects.filter( Q(batch_no=str(request.POST['invoice_search_key'])) |  Q(invoice_number=str(request.POST['invoice_search_key']))
                
            | Q(total_bill_amount=str(request.POST['invoice_search_key'])) | Q(cust_alph=str(request.POST['invoice_search_key'])))

        invoice_arr = []

        for invoice in selected_invoices:

            invoice_arr.append(model_to_dict(invoice))

        request.session['invoice_arr'] = json.dumps(invoice_arr)

        request.session['invoice_search_key'] = str(request.POST['invoice_search_key'])

        request.session['invoice_view'] = 'detail'
            
        return redirect('/invoice_board')



    if 'kt_tc_batch_no' in request.POST:

        selected_kt_tcs = TimeCard_KT.objects.filter(batch_no=str(request.POST['kt_tc_batch_no']))

        kt_tc_arr = []

        for kt_tc in selected_kt_tcs:

            kt_tc_arr.append(model_to_dict(kt_tc))

        request.session['timecard_kt_arr'] = json.dumps(kt_tc_arr)

        request.session['kt_tc_batch_no'] = str(request.POST['kt_tc_batch_no'])

        request.session['kt_tc_search_key'] = ''

        request.session['kt_tc_view'] = 'detail'
            
        return redirect('/timecard_board')


    if 'kt_tc_search_key' in request.POST:

        selected_kt_tcs = []

        if str(request.POST['kt_tc_search_key']) == '' :

            selected_kt_tcs = TimeCard_KT.objects.all()

        else:

            selected_kt_tcs = TimeCard_KT.objects.filter( Q(batch_no=str(request.POST['kt_tc_search_key'])) |  Q(time_card_id=str(request.POST['kt_tc_search_key'])))


        kt_tc_arr = []

        for kt_tc in selected_kt_tcs:

            kt_tc_arr.append(model_to_dict(kt_tc))

        request.session['timecard_kt_arr'] = json.dumps(kt_tc_arr)

        request.session['kt_tc_search_key'] = str(request.POST['kt_tc_search_key'])

        request.session['kt_tc_view'] = 'detail'
            
        return redirect('/timecard_board')


    if 'hb_tc_batch_no' in request.POST:

        selected_hb_tcs = TimeCard_HB.objects.filter(batch_no=str(request.POST['hb_tc_batch_no']))

        hb_tc_arr = []

        for hb_tc in selected_hb_tcs:

            hb_tc_arr.append(model_to_dict(hb_tc))

        request.session['timecard_hb_arr'] = json.dumps(hb_tc_arr)

        request.session['hb_tc_batch_no'] = str(request.POST['hb_tc_batch_no'])

        request.session['hb_tc_search_key'] = ''

        request.session['hb_tc_view'] = 'detail'
            
        return redirect('/timecard_hb_board')


    if 'hb_tc_search_key' in request.POST:

        selected_hb_tcs = []

        if str(request.POST['hb_tc_search_key']) == '' :

            selected_hb_tcs = TimeCard_HB.objects.all()

        else:

            selected_hb_tcs = TimeCard_HB.objects.filter( Q(batch_no=str(request.POST['hb_tc_search_key'])) |  Q(time_card_id=str(request.POST['hb_tc_search_key'])))

        hb_tc_arr = []

        for hb_tc in selected_hb_tcs:

            hb_tc_arr.append(model_to_dict(hb_tc))

        request.session['timecard_hb_arr'] = json.dumps(hb_tc_arr)

        request.session['hb_tc_search_key'] = str(request.POST['hb_tc_search_key'])

        request.session['hb_tc_view'] = 'detail'
            
        return redirect('/timecard_hb_board')


    if 'kt_rk_batch_no' in request.POST:

        selected_kt_rks = ReconKeys_KT.objects.filter(batch_no=str(request.POST['kt_rk_batch_no']))

        kt_rk_arr = []

        for kt_rk in selected_kt_rks:

            kt_rk_arr.append(model_to_dict(kt_rk))

        request.session['reconkeys_arr'] = json.dumps(kt_rk_arr)

        request.session['kt_rk_batch_no'] = str(request.POST['kt_rk_batch_no'])

        request.session['kt_rk_search_key'] = ''

        request.session['kt_rk_view'] = 'detail'
            
        return redirect('/reconkeys_board')


    if 'kt_rk_search_key' in request.POST:

        selected_kt_rks = []

        if str(request.POST['kt_rk_search_key']) == '' :

            selected_kt_rks = ReconKeys_KT.objects.all()

        else:

            selected_kt_rks = ReconKeys_KT.objects.filter( Q(batch_no=str(request.POST['kt_rk_search_key'])) |  Q( authorization=str(request.POST['kt_rk_search_key'])))

        kt_rk_arr = []

        for kt_rk in selected_kt_rks:

            kt_rk_arr.append(model_to_dict(kt_rk))

        request.session['reconkeys_arr'] = json.dumps(kt_rk_arr)

        request.session['kt_rk_search_key'] = str(request.POST['kt_rk_search_key'])

        request.session['kt_rk_view'] = 'detail'
            
        return redirect('/reconkeys_board')


    if 'hb_rk_batch_no' in request.POST:

        selected_hb_rks = ReconKeys_HB.objects.filter(batch_no=str(request.POST['hb_rk_batch_no']))

        hb_rk_arr = []

        for hb_rk in selected_hb_rks:

            hb_rk_arr.append(model_to_dict(hb_rk))

        request.session['reconkeys_hb_arr'] = json.dumps(hb_rk_arr)

        request.session['hb_rk_batch_no'] = str(request.POST['hb_rk_batch_no'])

        request.session['hb_rk_search_key'] = ''

        request.session['hb_rk_view'] = 'detail'
            
        return redirect('/reconkeys_hb_board')


    if 'hb_rk_search_key' in request.POST:

        selected_hb_rks = []

        if str(request.POST['hb_rk_search_key']) == '' :

            selected_hb_rks = ReconKeys_HB.objects.all()

        else:

            selected_hb_rks = ReconKeys_HB.objects.filter( Q(batch_no=str(request.POST['hb_rk_search_key'])) |  Q( row_id=str(request.POST['hb_rk_search_key'])))

        hb_rk_arr = []

        for hb_rk in selected_hb_rks:

            hb_rk_arr.append(model_to_dict(hb_rk))

        request.session['reconkeys_hb_arr'] = json.dumps(hb_rk_arr)

        request.session['hb_rk_search_key'] = str(request.POST['hb_rk_search_key'])

        request.session['hb_rk_view'] = 'detail'
            
        return redirect('/reconkeys_hb_board')



    if 'payment_batch_no' in request.POST:

        selected_payments = Payment.objects.filter(batch_no=str(request.POST['payment_batch_no']))

        input_payment_arr = []

        for payment in selected_payments:

            input_payment_arr.append(model_to_dict(payment))

        request.session['input_payment_arr'] = json.dumps(input_payment_arr)

        request.session['payment_batch_no'] = str(request.POST['payment_batch_no'])

        request.session['payment_search_key'] = ''

        request.session['payment_view'] = 'detail'
            
        return redirect('/payment_board')


    if 'payment_search_key' in request.POST:

        selected_payments = []

        if str(request.POST['payment_search_key']) == '' :

            selected_payments = Payment.objects.all()

        else:

            selected_payments = Payment.objects.filter( Q(batch_no=str(request.POST['payment_search_key'])) |  Q( payment_id=str(request.POST['payment_search_key'])))

        
        input_payment_arr = []

        for payment in selected_payments:

            input_payment_arr.append(model_to_dict(payment))

        request.session['input_payment_arr'] = json.dumps(input_payment_arr)

        request.session['payment_search_key'] = str(request.POST['payment_search_key'])

        request.session['payment_view'] = 'detail'
            
        return redirect('/payment_board')


    if 'cash_post_search_key' in request.POST:

        cash_response = str(request.POST['cash_post_search_key'])

        selected_cash_posts = []

        if cash_response == '' :

            selected_cash_posts = Cash_Post.objects.all()
        else :

            selected_cash_posts = Cash_Post.objects.filter( Q(cash_post_id=cash_response) | Q(recon_key=cash_response) )

        matching_by_recon_key_arr = []

        for matching in selected_cash_posts:

            matching_by_recon_key_arr.append(model_to_dict(matching))


        request.session['cash_post_search_key'] = str(request.POST['cash_post_search_key'])
            
        # return redirect('/matching_by_recon_key')

        new_matching_arr = []

        imported_payment_arr = json.loads(request.session.get('imported_payment_arr', '[]'))

        imported_invoice_arr = json.loads(request.session.get('imported_invoice_arr', '[]'))

        input_payment_arr = json.loads(request.session.get('input_payment_arr', '[]'))

        for matching in matching_by_recon_key_arr:

            child_invoice = []

            child_payment = []

            child_ori_payment = []

            multiple = []

            for invoice in imported_invoice_arr:            

                if matching['recon_key'] in invoice['recon_key']:

                    child_invoice.append(invoice)

                    multiple = invoice['recon_key']

            for recon in multiple:

                for payment in imported_payment_arr:

                    if recon in payment['recon_key']:

                        child_payment.append(payment)

                for payment in input_payment_arr:

                    if recon in payment['recon_key'] and payment not in child_ori_payment:

                        child_ori_payment.append(payment)


            matching['child_invoice'] = child_invoice

            matching['child_payment'] = child_payment

            matching['child_ori_payment'] = child_ori_payment

            new_matching_arr.append(matching)


        request.session['matching_by_recon_key_arr'] = json.dumps(new_matching_arr)


        return render(request, 'report/matching_by_recon_key.html',
            {

                'matching_by_recon_key_arr' : json.loads(request.session.get('matching_by_recon_key_arr', '[]')),

                'cash_post_batch_no_arr' : request.session.get('cash_post_batch_no_arr', '[]'),

                'cash_post_search_key' : request.session.get('cash_post_search_key', ''),

            })


def cash_post(request):

    cash_post_arr = json.loads(request.session.get('matching_by_recon_key_arr', '[]'))

    for cash_post in cash_post_arr:

        check = Cash_Post.objects.filter(recon_key = cash_post['recon_key'])

        if 'child_invoice' in cash_post:

            del cash_post['child_invoice']

        if 'child_payment' in cash_post:

            del cash_post['child_payment']

        if 'child_ori_payment' in cash_post:

            del cash_post['child_ori_payment']

        if 'comment' in cash_post:

            del cash_post['comment']

        if len(check) == 0:

            cash_post_model = Cash_Post(**cash_post)

            cash_post_model.save()

        else :

            check.update(**cash_post)

    return render(request, 'report/matching_by_recon_key.html',
        {

            'matching_by_recon_key_arr' : json.loads(request.session.get('matching_by_recon_key_arr', '[]')),

            'cash_post_batch_no_arr' : request.session.get('cash_post_batch_no_arr', '[]'),

            'selected_cash_post_id' : request.session.get('selected_cash_post_id', ''),

        })

    # return redirect('/matching_by_recon_key')


def add_comment(request):

    now = datetime.datetime.now()

        # data_table = Template("""
        #     <li class="media">
        #     <a class="pull-left" href="javascript:;">
        #         <img alt="" class="img-circle" src="/static/pages/img/avatars/team16.jpg" style="width: 40px;"></a>
        #     <div class="media-body todo-comment">
        #         <p class="todo-comment-head">
        #              <span class="property-name">{{ user }}</span> &nbsp;
        #             <span class="todo-comment-date"> {{ date }}</span>
        #         </p>
        #         <p class="todo-text-color"> {{ comment }}</p>
        #     </div>
        #     </li>
        #         """)

        # html = data_table.render(Context({

        #     'comment': request.GET['comment'], 

        #     'date' : '{0.month}/{0.day}/{0.year} at {0.hour}:{0.minute}'.format(now),

        #     'user' : request.session['user']['username']

        #     }))


    data = {

        'content' : request.GET['comment'],

        'recon_key' : request.GET['recon_key'],

        'posted_timestamp' : request.GET['posted_timestamp'],

        'posted_by' : request.GET['posted_by']
    }

    comment_model = Comment(**data)

    comment_model.save()

    return JsonResponse({ 'status' : 'success' })


def set_collection_status(request):

    check = Cash_Post.objects.filter(recon_key = request.GET['recon_key'])

    data = model_to_dict(check[0])

    data['collection_status'] = request.GET['collection_status']

    check.update(**data)

    return HttpResponse('success')


def main(request):
    # clients = Client.objects.all().order_by("-created_on")
    return render(request, 'black.html')
    # return render_to_response('blank.html', locals(), context_instance=RequestContext(request))

def home(request):
    # for key in list(request.session.keys()):

    #     if key !='user' :

    #         del request.session[key]

    return render(request, 'home.html')
    

def invoice_board(request):

    inv_bhs = Invoice.objects.all().values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    invoice_batch_no_arr = []

    for batch in inv_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }


        invoice_batch_no_arr.append(data)

    request.session['invoice_batch_no_arr'] = invoice_batch_no_arr


    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = Invoice.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-INV-0000'

        num = str(int(latest_batch.split('-')[2])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:2])+'-'+num

        request.session['invoice_batch_no'] = batch_no

        invoice_arr = []

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('data/csv/invoice/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading Invoice Data ...')

        input_invoice_data_sheet = workbook[workbook.sheetnames[0]]

        input_invoice_data_col_max_count = 0

        invoice_row = 1

        invoice_header = [
            'company', 
            'branch_id', 
            'employee_short_name', 
            'cust_alph', 
            'customer_name', 
            'invoice_number', 
            'straight_time_hours', 
            'hourly_bill_rate', 
            'overtime_hours', 
            'sp_bill', 
            'total_bill_amount', 
            'invoice_date', 
            'week_ending_date', 
            'invoice_reference', 
            'workers_comp_code', 
            'job_description', 
            'customer_po', 
            'b_w_location_code', 
            'address_1', 
            'spcl_paybill_de_code', 
            'special_paybill_qty', 
            'special_bill_rate'
         ]

        now = datetime.datetime.now()

        for invoice in input_invoice_data_sheet.rows:

            try:

                data = {

                    'batch_no' : batch_no,

                    'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                }

                for invoice_col in range(0, len(invoice_header)) :

                    data[invoice_header[invoice_col]] = str(invoice[invoice_col].value)

                if invoice_row > 1:

                    items = data['invoice_reference'].split('@')

                    try:

                        data['system'] = items[0]

                        data['license'] = items[1]

                        data['age'] = items[2]

                        data['discipline'] = items[3]

                        data['modes'] = items[4]

                        data['tc'] = items[5]

                    except:

                        pass

                    invoice_arr.append(data)

                    check = Invoice.objects.filter(invoice_reference=data['invoice_reference'])

                    if len(check) == 0:

                        invoice_model = Invoice(**data)

                        invoice_model.save()

                    else :

                        data['batch_no'] = check[0].batch_no

                        check.update(**data)


                invoice_row += 1

            except:

                pass

        request.session['invoice_arr'] = json.dumps(invoice_arr)

        request.session['invoice_search_key'] = ''

        # data_table = Template("""
        #     """)
        # html = data_table.render(Context({'invoice_arr': json.loads(request.session.get('invoice_arr', '[]'))}))

        return HttpResponse("success")


    else:

        return render(request, 'invoice/index.html',
            {

                'invoice_arr' : json.loads(request.session.get('invoice_arr', '[]')),

                'invoice_batch_no_arr' : request.session.get('invoice_batch_no_arr', '[]'),

                'invoice_batch_no' : request.session.get('invoice_batch_no', ''),

                'view' : request.session.get('invoice_view', 'batch'),

                'invoice_search_key' : request.session.get('invoice_search_key', '')

            })


def invoice_simple_board(request):

    if request.method == 'POST' and request.FILES:

        invoice_arr = []

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('uploaded_files/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading Invoice Data ...')

        invoice_data_sheet = workbook[workbook.sheetnames[0]]        

        invoice_header = [
           'invoice',
           'invoice_amount',
           'recon_key'
         ]        

        input_invoice_arr = []

        invoice_row = 1

        for invoice in invoice_data_sheet.rows:

            try:

                amount = 0

                recon_key = []

                if invoice[1].value != None and invoice[1].value != '':

                    amount = str(invoice[1].value)

                for cnt in range(2, len(invoice) ) :

                    if invoice[cnt].value != None and invoice[cnt].value != '':

                        recon_key.append( str(invoice[cnt].value).replace(' ', '').replace('-', '').strip() )

                item = {

                    'invoice' : str(invoice[0].value),

                    'invoice_amount' : amount,

                    'recon_key' : recon_key,

                }

                if invoice_row > 1:

                    input_invoice_arr.append( item )

                invoice_row += 1

            except :

                pass

        # pdb.set_trace()

        request.session['input_invoice_arr'] = json.dumps(input_invoice_arr)

        # data_table = Template("""
        #     """)
        # html = data_table.render(Context({'invoice_arr': json.loads(request.session.get('invoice_arr', '[]'))}))

        return HttpResponse("success")

    else:
        
        return render(request, 'invoice/invoice_simple.html',
            {

                'invoice_arr' : json.loads(request.session.get('input_invoice_arr', '[]')),

            })


def timecard_board(request):

    kt_tc_bhs = TimeCard_KT.objects.values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    kt_tc_batch_no_arr = []

    for batch in kt_tc_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }

        kt_tc_batch_no_arr.append(data)

    request.session['kt_tc_batch_no_arr'] = kt_tc_batch_no_arr

    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = TimeCard_KT.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-KT-TC-0000'

        num = str(int(latest_batch.split('-')[3])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:3])+'-'+num

        request.session['kt_tc_batch_no'] = batch_no

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('data/csv/timecard/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading KT Time Card Time Card Data ...')

        kt_tc_raw_sheet = workbook[workbook.sheetnames[0]]

        timecard_kt_arr = []

        kt_tc_row = 1

        timecard_header = [
            'time_card_id',
            'authorization',
            'bill_mode',
            'bill_rate',
            'billable',
            'case_manager_first_name',
            'case_manager_last_name',
            'checkin_time',
            'checkout_time',
            'client_first_name',
            'client_last_name',
            'client_dob',
            'clinician_first_name',
            'clinician_last_name',
            'clinician_id',
            'clinician_payroll_group',
            'clinician_payroll_id',
            'clinician_payroll_rule',
            'clinician_skills',
            'clinician_ssn',
            'cosign_staff_first_name',
            'cosign_staff_last_name',
            'edited_hours',
            'episode_duration',
            'episode_number',
            'episode_start_date',
            'episode_end_date',
            'has_split_schedules',
            'hcpcs_code',
            'location',
            'marital_status',
            'medicaid',
            'patient_id',
            'payroll',
            'payroll_mode',
            'payroll_rate',
            'payrolled',
            'schedule_status',
            'units',
            'visit_date',
            'client_lob',
            'insurance_id',
            'payer',
            'clinician_discipline',
            'service',
            'employment_type',
            'invoice_no',
            'miles',
            'visit_id',
            'payroll_batch_id',
            'authorization_2',
            'bill_mode_2',
            'shift',
            'uid'
        ]

        now = datetime.datetime.now()

        for kt_tc_raw in kt_tc_raw_sheet.rows:

            try:

                data = {

                    'batch_no' : batch_no,

                    'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                } 

                for kt_tc_col in range(0, len(timecard_header)-1 ) :

                    data[timecard_header[kt_tc_col]] = kt_tc_raw[kt_tc_col].value
                    

                temp = datetime.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!

                date = ''

                try:

                    delta = data['visit_date'] - temp

                    date = str(int(float(delta.days) + (float(delta.seconds) / 86400)))

                except : 

                    date = '0'
                
                time = ''

                try:

                    time = float(float(data['checkin_time'].hour)+float(data['checkin_time'].minute)/60+float(data['checkin_time'].second)/3600)/24     

                    if time == 0:

                        time = '0'

                except :

                    time = '0'

                uid = str(data['clinician_ssn']) + '@' + str(data['location']) + '@' + str(data['patient_id']) + '@' + date + '@' + str(time) + '@' + str(data['service'])

                data['uid'] = uid.strip()

                for header in timecard_header:

                    data[header] = str(data[header])

                if kt_tc_row > 1 and data['time_card_id'] != '':

                    timecard_kt_arr.append(data)

                    check = TimeCard_KT.objects.filter(time_card_id=data['time_card_id'])

                    if len(check) == 0:

                        timecard_kt_model = TimeCard_KT(**data)

                        timecard_kt_model.save()

                    else :

                        data['batch_no'] = check[0].batch_no

                        check.update(**data)

                kt_tc_row += 1

            except:

                pass


        request.session['timecard_kt_arr'] = json.dumps(timecard_kt_arr)

        request.session['kt_tc_search_key'] = ''

        # data_table = Template("""
        #     """)
        # html = data_table.render(Context({'timecard_kt_arr': json.loads(request.session.get('timecard_kt_arr', '[]'))}))

        return HttpResponse("success")

    else:
        
        return render(request, 'timecard/timecard_kt.html',
            {

                'timecard_kt_arr' : json.loads(request.session.get('timecard_kt_arr', '[]')),

                'kt_tc_batch_no_arr' : request.session.get('kt_tc_batch_no_arr', '[]'),

                'kt_tc_search_key' : request.session.get('kt_tc_search_key', ''),

                'view' : request.session.get('kt_tc_view', 'batch')

            })


def timecard_hb_board(request):

    hb_tc_bhs = TimeCard_HB.objects.values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    hb_tc_batch_no_arr = []

    for batch in hb_tc_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }

        hb_tc_batch_no_arr.append(data)

    request.session['hb_tc_batch_no_arr'] = hb_tc_batch_no_arr


    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = TimeCard_HB.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-HB-TC-0000'

        num = str(int(latest_batch.split('-')[3])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:3])+'-'+num

        request.session['hb_tc_batch_no'] = batch_no

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('data/csv/timecard/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading HB Time Card Time Card Data ...')

        hb_tc_raw_sheet = workbook[workbook.sheetnames[0]]

        timecard_hb_arr = []

        hb_tc_row = 1

        timecard_header = [
            'time_card_id',
            'batch',
            'worker_name',
            'primary_job_desc',
            'standard_pay_method',
            'payroll_branch',
            'client_branch',
            'client_name',
            'client_state',
            'service_date',
            'shift',
            'drive_start_time',
            'in_home_nva_start_time',
            'in_home_nva_end_time',
            'service_type',
            'rate',
            'service_code_rate_type',
            'total_time',
            'qty',
            'payroll_transmittal_group',
            'prod_points',
            'travel_am',
            'travel_tf',
            'travel_cv',
            'mileage_payment_method',
            'service_payable',
            'service_payment_method',
            'worker_id',
            'full_time',
            'default_milleage_payment_method',
            'worker_class',
            'worker_category',
            'payroll_department',
            'payroll_number',
            'expected_prod_points',
            'team',
            'municipality_code',
            'date_of_hire',
            'uid'
        ]

        now = datetime.datetime.now()

        for hb_tc_raw in hb_tc_raw_sheet.rows:

            try:

                data = {

                    'batch_no' : batch_no,

                    'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                } 

                for hb_tc_col in range(0, len(timecard_header)-1 ) :

                    data[timecard_header[hb_tc_col]] = hb_tc_raw[hb_tc_col].value

                temp = datetime.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!

                date = ''

                try:

                    delta = data['service_date'] - temp

                    date = str(int(float(delta.days) + (float(delta.seconds) / 86400)))

                except : 

                    date = '0'

                uid = str(data['client_branch']) + '@' + date + '@' + str(data['service_type']) + '@' + str(data['client_name']).replace(' ', '').replace('.', '').strip()

                data['uid'] = uid.strip()

                for header in timecard_header:

                    data[header] = str(data[header])

                if hb_tc_row > 1 and data['time_card_id'] != '' :

                    timecard_hb_arr.append(data)

                    check = TimeCard_HB.objects.filter(time_card_id=data['time_card_id'])

                    if len(check) == 0:

                        timecard_hb_model = TimeCard_HB(**data)

                        timecard_hb_model.save()

                    else :

                        data['batch_no'] = check[0].batch_no
                        
                        check.update(**data)

                hb_tc_row += 1

            except:

                pass

        request.session['timecard_hb_arr'] = json.dumps(timecard_hb_arr)

        request.session['hb_tc_search_key'] = ''
        # data_table = Template("""
        #     """)
        # html = data_table.render(Context({'timecard_kt_arr': json.loads(request.session.get('timecard_kt_arr', '[]'))}))

        return HttpResponse("success")

    else:
        
        return render(request, 'timecard/timecard_hb.html',
            {

                'timecard_hb_arr' : json.loads(request.session.get('timecard_hb_arr', '[]')),

                'hb_tc_batch_no_arr' : request.session.get('hb_tc_batch_no_arr', '[]'),

                'hb_tc_search_key' : request.session.get('hb_tc_search_key', ''),

                'view' : request.session.get('hb_tc_view', 'batch')

            })


def reconkeys_board(request):

    kt_rk_bhs = ReconKeys_KT.objects.values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    kt_rk_batch_no_arr = []

    for batch in kt_rk_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }

        kt_rk_batch_no_arr.append(data)

    request.session['kt_rk_batch_no_arr'] = kt_rk_batch_no_arr


    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = ReconKeys_KT.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-KT-RK-0000'            

        num = str(int(latest_batch.split('-')[3])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:3])+'-'+num

        request.session['kt_rk_batch_no'] = batch_no

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('data/csv/reconkey/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading KT RK Data ...')

        kt_key_raw_sheet = workbook[workbook.sheetnames[0]]

        reconkeys_arr = []

        reconkeys_header = [
            'authorization',
            'bill_mode',
            'bill_rate',
            'billable',
            'case_manager_first_name',
            'case_manager_last_name',
            'checkin_time',
            'checkout_time',
            'client_first_name',
            'client_last_name',
            'client_dob',
            'clinician_first_name',
            'clinician_last_name',
            'clinician_id',
            'clinician_payroll_group',
            'clinician_payroll_id', 
            'clinician_payroll_rule',
            'clinician_skills',
            'clinician_ssn',
            'cosign_staff_first_name',
            'cosign_staff_last_name',
            'edited_hours',
            'episode_duration',
            'episode_number',
            'episode_start_date',
            'episode_end_date',
            'has_split_schedules',
            'hcpcs_code',
            'location',
            'material_status',
            'medicaid',
            'patient_id',
            'payroll',  
            'payroll_mode',
            'payroll_rate',
            'payrolled',
            'schedule_status',
            'units',
            'visit_date',
            'client_lob',
            'insurance_id',
            'payer',
            'clinician_discipline',
            'service',
            'employment_type',
            'invoice_no',
            'miles',
            'visit_id',
            'payroll_batch_id',
            'shift',
            'visit_split_id',
            'key_id'
        ]

        now = datetime.datetime.now()

        kt_key_row = 1

        for kt_key_raw in kt_key_raw_sheet.rows:

            try:

                data = {

                    'batch_no' : batch_no,

                    'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                } 

                for kt_key_col in range(0, len(reconkeys_header)-1 ) :

                    data[reconkeys_header[kt_key_col]] = kt_key_raw[kt_key_col].value

                temp = datetime.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!

                date = ''

                try:

                    delta = data['visit_date'] - temp

                    date = str(int(float(delta.days) + (float(delta.seconds) / 86400)))

                except : 

                    date = '0'

                key_id = str(data['clinician_ssn']) + '@' + str(data['location']) + '@' + str(data['patient_id']) + '@' + date + '@' + str(data['checkin_time']) + '@' + str(data['service'])

                data['key_id'] = key_id.strip()

                data['invoice_no'] = str(data['invoice_no']).replace(' ', '').replace('-', '').strip()

                for header in reconkeys_header:

                    data[header] = str(data[header])

                if kt_key_row > 1 and data['insurance_id'] != '':

                    reconkeys_arr.append(data)

                    check = ReconKeys_KT.objects.filter(key_id=data['key_id'])

                    if len(check) == 0:

                        reconkeys_kt_model = ReconKeys_KT(**data)

                        reconkeys_kt_model.save()

                    else :

                        data['batch_no'] = check[0].batch_no

                        check.update(**data)


                kt_key_row += 1


            except :

                pass

        request.session['reconkeys_arr'] = json.dumps(reconkeys_arr)

        request.session['kt_rk_search_key'] = ''

        # data_table = Template("""
        #     """)

        # html = data_table.render(Context({'reconkeys_arr': json.loads(request.session.get('reconkeys_arr', '[]'))}))

        if len(reconkeys_arr) == 0:

            request.session['msg'] = 'Invalid Format'

        else :

            request.session['msg'] = ''            

        return HttpResponse("success")

    else:
        
        return render(request, 'reconkeys/reconkeys_kt.html',
            {

                'reconkeys_arr' : json.loads(request.session.get('reconkeys_arr', '[]')),

                'kt_rk_batch_no_arr' : request.session.get('kt_rk_batch_no_arr', '[]'),

                'kt_rk_search_key' : request.session.get('kt_rk_search_key', ''),

                'view' : request.session.get('kt_rk_view', 'batch')

            })


def reconkeys_hb_board(request):

    hb_rk_bhs = ReconKeys_HB.objects.values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    hb_rk_batch_no_arr = []

    for batch in hb_rk_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }

        hb_rk_batch_no_arr.append(data)

    request.session['hb_rk_batch_no_arr'] = hb_rk_batch_no_arr 


    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = ReconKeys_HB.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-HB-RK-0000'            

        num = str(int(latest_batch.split('-')[3])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:3])+'-'+num

        request.session['hb_rk_batch_no'] = batch_no

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        filename = fs.save('data/csv/reconkey/'+myfile.name, myfile)

        uploaded_file_url = fs.url(filename)

        workbook = load_workbook( myfile )

        print(' Loading HB RK Data ...')

        hb_key_raw_sheet = workbook[workbook.sheetnames[0]]

        reconkeys_hb_arr = []

        reconkeys_header = [
            'row_id',
            'branch',
            'funding_source',
            'post_date',
            'date_created',
            'client_facility_name',
            'shift_date',
            'invoice',
            'description',
            'rev_code',
            'hcpcs_cpt_code',
            'quantity_units',
            'orig_charge',
            'adjustments',
            'amount',
            'balance',
            'key_id'
        ]

        hb_key_row = 1

        now = datetime.datetime.now()

        for hb_key_raw in hb_key_raw_sheet.rows:

            try:

                data = {

                    'batch_no' : batch_no,

                    'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                } 

                for hb_key_col in range(0, len(reconkeys_header)-1 ) :

                    data[reconkeys_header[hb_key_col]] = hb_key_raw[hb_key_col].value

                temp = datetime.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!

                date = ''

                try:

                    delta = data['shift_date'] - temp

                    date = str(int(float(delta.days) + (float(delta.seconds) / 86400)))

                except : 

                    date = str(data['shift_date'])

                key_id = str(data['branch']) + '@' + date + '@' + str(data['description']) + '@' + str(data['client_facility_name']).replace(' ', '').replace('.', '').strip()

                data['key_id'] = key_id.strip()

                data['invoice'] = str(data['invoice']).replace(' ', '').replace('-', '').strip()

                for header in reconkeys_header:

                    data[header] = str(data[header])

                if hb_key_row > 1 and data['branch'] != '':

                    reconkeys_hb_arr.append(data)

                    check = ReconKeys_HB.objects.filter(key_id=data['key_id'])

                    if len(check) == 0:

                        reconkeys_hb_model = ReconKeys_HB(**data)

                        reconkeys_hb_model.save()

                    else :

                        data['batch_no'] = check[0].batch_no

                        check.update(**data)

                hb_key_row += 1

            except :

                pdb.set_trace()

        request.session['reconkeys_hb_arr'] = json.dumps(reconkeys_hb_arr)

        request.session['hb_rk_search_key'] = ''


        # data_table = Template("""
        #     """)

        # html = data_table.render(Context({'reconkeys_arr': json.loads(request.session.get('reconkeys_arr', '[]'))}))

        if len(reconkeys_hb_arr) == 0:

            request.session['msg'] = 'Invalid Format'

        else :

            request.session['msg'] = ''            

        return HttpResponse("success")

    else:
        
        return render(request, 'reconkeys/reconkeys_hb.html',
            {

                'reconkeys_hb_arr' : json.loads(request.session.get('reconkeys_hb_arr', '[]')),

                'hb_rk_batch_no_arr' : request.session.get('hb_rk_batch_no_arr', '[]'),

                'hb_rk_search_key' : request.session.get('hb_rk_search_key', ''),

                'view' : request.session.get('hb_rk_view', 'batch')

            })


def payment_board(request):

    payment_bhs = Payment.objects.values('batch_no', 'uploaded_date').annotate(count=Count('batch_no'))

    payment_batch_no_arr = []

    for batch in payment_bhs:

        data = {

            'batch_no' : batch['batch_no'],

            'count': batch['count'],

            'uploaded_date' : batch['uploaded_date']
        }

        payment_batch_no_arr.append(data)


    request.session['payment_batch_no_arr'] = payment_batch_no_arr 

    if request.method == 'POST' and request.FILES:

        latest_batch = ''

        try:

            latest_batch = Payment.objects.latest('batch_no').batch_no

        except:

            latest_batch = 'BH-PAY-0000'

        num = str(int(latest_batch.split('-')[2])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        batch_no = '-'.join(latest_batch.split('-')[:2])+'-'+num

        request.session['payment_batch_no'] = batch_no


        latest_payment_id = ''

        try:

            latest_payment_id = Payment.objects.latest('payment_id').payment_id

        except:

            latest_payment_id = 'EOB0000'

        num = str(int(latest_payment_id[3:])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        start_payment_id = latest_payment_id[:3]+num

        payment_arr = []

        myfile = request.FILES['file']

        fs = FileSystemStorage()

        if '.pdf' in myfile.name.lower():

            filename = fs.save('data/pdf/'+myfile.name, myfile)

            items = []
            items.append(filename)

            payment_arr = parse_superior_template(request, items)


        else:

            filename = fs.save('data/csv/payment/'+myfile.name, myfile)

            uploaded_file_url = fs.url(filename)

            workbook = load_workbook( myfile )

            input_payment_data_sheet = workbook[workbook.sheetnames[0]]

            payment_header = [

                'check',

                'recon_key',

                'check_amount'

            ]

            payment_row = 1

            now = datetime.datetime.now()

            payment_id = start_payment_id

            for payment in input_payment_data_sheet.rows:

                try:

                    if payment_row > 1:

                        data = {

                            'batch_no' : batch_no,

                            'payment_id' : payment_id,

                            'check' : str(payment[0].value),

                            'recon_key' : str(payment[1].value).replace(' ', '').replace('-', '').strip(),

                            'check_amount' : str(payment[2].value).replace('$',''),

                            'uploaded_date' : '{0.month}/{0.day}/{0.year}'.format(now)

                        }

                        payment_arr.append( data )

                        num = str(int(payment_id[3:])+1)

                        if len(num) < 4:

                            for ind in range(0, 4-len(num)):

                                num = '0'+num

                        payment_id = payment_id[:3]+num

                        check = Payment.objects.filter(check=data['check'], recon_key=data['recon_key'], check_amount=data['check_amount'])

                        if len(check) == 0:

                            payment_model = Payment(**data)

                            payment_model.save()

                        else :

                            data['batch_no'] = check[0].batch_no

                            data['cash_post_id_memo'] = check[0].cash_post_id_memo

                            check.update(**data)

                    payment_row += 1


                except:

                    pass

        temp_data = payment_arr

        request.session['input_payment_arr'] = json.dumps(payment_arr)

        request.session['payment_search_key'] = ''

        # data_table = Template("""
        # """)
        # html = data_table.render(Context({'payment_arr': json.loads(request.session.get('payment_arr', '[]'))}))

        return HttpResponse("success")

    else:
        
        return render(request, 'payment/index.html',

            {

                'input_payment_arr' : json.loads(request.session.get('input_payment_arr', '[]')),

                'payment_batch_no_arr' : request.session.get('payment_batch_no_arr', '[]'),

                'payment_search_key' : request.session.get('payment_search_key', ''),

                'view' : request.session.get('payment_view', 'batch')

            })


def generated_invoice(request):

    return render(request, 'report/generated_invoice.html',
        {

            'input_invoice_arr' : json.loads(request.session.get('input_invoice_arr', '[]')),

        })


def imported_payment(request):

    return render(request, 'report/imported_payment.html',
        {

            'imported_payment_arr' : json.loads(request.session.get('imported_payment_arr', '[]')),

        })


def unused_payment(request):

    return render(request, 'report/unused_payment.html',
        {

            'unused_payment_arr' : json.loads(request.session.get('unused_payment_arr', '[]')),

        })


def unused_invoice(request):

    return render(request, 'report/unused_invoice.html',
        {

            'unused_invoice_arr' : json.loads(request.session.get('unused_invoice_arr', '[]')),

        })

def default_module(request):

    cash_post_bhs = Cash_Post.objects.all().values('cash_post_id', 'posted_date').annotate(count=Count('cash_post_id'))

    cash_post_batch_no_arr = []

    for batch in cash_post_bhs:

        data = {

            'cash_post_id' : batch['cash_post_id'],

            'count': batch['count'],

            'posted_date' : batch['posted_date']
        }


        cash_post_batch_no_arr.append(data)

    request.session['cash_post_batch_no_arr'] = cash_post_batch_no_arr


    input_invoice_arr = []

    max_recon_count = 0

    invoices_with_rks_count = 0

    invoices_without_rks_count = 0


    used_kt_tc_data = []

    unused_kt_tc_data = []


    used_kt_key_data = []

    unused_kt_key_data = []


    used_hb_tc_data = []

    unused_hb_tc_data = []

    used_hb_key_data = []

    unused_hb_key_data = []

    now = datetime.datetime.now()

    today = '{0.month}/{0.day}/{0.year}'.format(now)

    invoice_db = Invoice.objects.all()

    invoice_data_arr = []

    for invoice in invoice_db:

        invoice_data_arr.append(model_to_dict(invoice))


    timecard_kt_db = TimeCard_KT.objects.all()

    kt_tc_raw_arr = []

    for timecard_kt in timecard_kt_db:

        kt_tc_raw_arr.append(model_to_dict(timecard_kt))


    timecard_hb_db = TimeCard_HB.objects.all()

    hb_tc_raw_arr = []

    for timecard_hb in timecard_hb_db:

        hb_tc_raw_arr.append(model_to_dict(timecard_hb))   


    reconkeys_kt_db = ReconKeys_KT.objects.all()

    kt_key_raw_arr = []

    for reconkeys_kt in reconkeys_kt_db:

        kt_key_raw_arr.append(model_to_dict(reconkeys_kt))   


    reconkeys_hb_db = ReconKeys_HB.objects.all()

    hb_key_raw_arr = []

    for reconkeys_hb in reconkeys_hb_db:

        hb_key_raw_arr.append(model_to_dict(reconkeys_hb))

    # payment_db = Payment.objects.filter(uploaded_date__icontains=today)

    payment_db = Payment.objects.all()

    input_payment_arr = []

    for payment in payment_db:

        input_payment_arr.append(model_to_dict(payment))

    for invoice_data in invoice_data_arr:

        try:

            res = invoice_data

            res['invoice'] = invoice_data['invoice_number']

            res['recon_key'] = []

            res['invoice_amount'] = invoice_data['total_bill_amount']

            res['timecard_kt'] = []

            res['timecard_hb'] = []

            recon_count = 0

            for kt_tc_raw in kt_tc_raw_arr : 

                try:

                    if str(invoice_data['system']) + str(invoice_data['tc']) == str(kt_tc_raw['time_card_id']):


                        for kt_key_raw in kt_key_raw_arr : 

                            try:

                                if kt_tc_raw['uid'] == kt_key_raw['key_id'] and kt_key_raw['invoice_no'] not in res['recon_key']: 

                                    # pdb.set_trace()

                                    res['recon_key'].append(kt_key_raw['invoice_no'])

                                    if kt_key_raw not in used_kt_key_data:

                                        used_kt_key_data.append(kt_key_raw)

                                    if kt_tc_raw not in used_kt_tc_data:

                                        used_kt_tc_data.append(kt_tc_raw)

                                    res['timecard_kt'].append(kt_tc_raw)

                                    recon_count += 1

                            except : 

                                pdb.set_trace()

                except :

                    pass

            for hb_tc_raw in hb_tc_raw_arr : 

                try:

                    if str(invoice_data['system']) + str(invoice_data['tc']) == str(hb_tc_raw['time_card_id']) :

                        for hb_key_raw in hb_key_raw_arr : 

                            try:

                                if hb_tc_raw['uid'] == hb_key_raw['key_id'] and hb_key_raw['invoice'] not in res['recon_key']: 

                                    res['recon_key'].append(hb_key_raw['invoice'])

                                    if hb_key_raw not in used_hb_key_data:

                                        used_hb_key_data.append(hb_key_raw)

                                    if hb_tc_raw not in used_hb_tc_data:

                                        used_hb_tc_data.append(hb_tc_raw)

                                    res['timecard_hb'].append(hb_tc_raw)

                                    recon_count += 1
                            except : 

                                pass

                except : 

                    pass

            if recon_count != 0:

                invoices_with_rks_count += 1

            else :

                invoices_without_rks_count += 1


            if max_recon_count < recon_count :

                max_recon_count = recon_count

            res['recon_key_view'] = ', '.join(res['recon_key'])


            input_invoice_arr.append(res)

        except :

            print('something went wrong')

            pdb.set_trace()

    request.session['input_invoice_arr'] = json.dumps(input_invoice_arr)

    request.session['input_payment_arr'] = json.dumps(input_payment_arr)

    # start cash posting.

    input_invoice_arr = json.loads(request.session.get('input_invoice_arr', '[]'))

    input_payment_arr = json.loads(request.session.get('input_payment_arr', '[]'))

    # calculating unused invoices 

    unused_invoice_arr = []

    imported_invoice_arr = []

    imported_payment_arr = []

    unused_payment_arr = []

    matching_by_recon_key_arr = []


    total_invoice_amount = 0

    reconned_invoice_amount = 0

    unused_invoice_amount = 0

    total_payment_amount = 0

    reconned_payment_amount = 0

    unused_payment_amount = 0

    balance_on_recon_invoice = 0


    try:

        for invoice in input_invoice_arr:

            flag = 0

            for payment in input_payment_arr:

                if payment['recon_key'] in invoice['recon_key']:

                    flag += 1

            if flag == 0:

                unused_invoice_arr.append(invoice)

                unused_invoice_amount += float(invoice['invoice_amount'])

            else :

                imported_invoice_arr.append(invoice)

                reconned_invoice_amount += float(invoice['invoice_amount'])

            total_invoice_amount += float(invoice['invoice_amount'])

        unique_recon_list = []

        request.session['imported_invoice_arr'] = json.dumps(imported_invoice_arr)

        for invoice in imported_invoice_arr:

            if invoice['recon_key'][0] not in unique_recon_list:

                unique_recon_list.append(invoice['recon_key'][0])


        request.session['unique_recon_list'] = json.dumps(unique_recon_list)

        # calculating imported payment and unused payment

        for payment in input_payment_arr:

            flag = 0

            sub_invoice_amt = 0

            for invoice in input_invoice_arr:

                if payment['recon_key'] in invoice['recon_key']:

                    sub_invoice_amt += float(invoice['invoice_amount'])

            for invoice in input_invoice_arr:

                if payment['recon_key'] in invoice['recon_key']:

                    imported_payment = {

                        'batch_no' : payment['batch_no'],

                        'payment_id' : payment['payment_id'],

                        'invoice' : invoice['invoice'],

                        'payment' : payment['check'],

                        'check_amount' : float(payment['check_amount']) * float(invoice['invoice_amount']) / sub_invoice_amt ,

                        'recon_key' : payment['recon_key'],

                        'cash_post_id_memo' : payment['cash_post_id_memo'],

                        'uploaded_date' : payment['uploaded_date']

                    }

                    imported_payment_arr.append(imported_payment)

                    flag += 1

            if flag == 0:

                unused_payment_arr.append(payment)

                unused_payment_amount += float(payment['check_amount'])

            else :

                reconned_payment_amount += float(payment['check_amount'])

            total_payment_amount += float(payment['check_amount'])


        balance_on_recon_invoice = reconned_invoice_amount - reconned_payment_amount

    except :

        pass


    request.session['imported_payment_arr'] = json.dumps(imported_payment_arr)

    request.session['unused_payment_arr'] = json.dumps(unused_payment_arr)

    request.session['unused_invoice_arr'] = json.dumps(unused_invoice_arr)

    request.session['total_invoice_amount'] = total_invoice_amount

    request.session['reconned_invoice_amount'] = reconned_invoice_amount

    request.session['unused_invoice_amount'] = unused_invoice_amount

    request.session['total_payment_amount'] = total_payment_amount

    request.session['reconned_payment_amount'] = reconned_payment_amount

    request.session['unused_payment_amount'] = unused_payment_amount

    request.session['balance_on_recon_invoice'] = balance_on_recon_invoice


def calculate(request):

    default_module(request)

    # input_invoice_arr = json.loads(request.session.get('input_invoice_arr', '[]'))

    try:

        # calculation matching by recon key data

        imported_payment_arr = json.loads(request.session.get('imported_payment_arr', '[]'))

        imported_invoice_arr = json.loads(request.session.get('imported_invoice_arr', '[]'))

        input_payment_arr = json.loads(request.session.get('input_payment_arr', '[]'))

        unique_recon_list = json.loads(request.session.get('unique_recon_list', '[]'))

        matching_by_recon_key_arr = []

        latest_cash_post_id = ''

        try:

            latest_cash_post_id = Cash_Post.objects.latest('cash_post_id').batch_no

        except:

            latest_cash_post_id = 'CP0000'

        num = str(int(latest_cash_post_id[2:])+1)

        if len(num) < 4:

            for ind in range(0, 4-len(num)):

                num = '0'+num

        latest_cash_post_id = latest_cash_post_id[:2] + num

        now = datetime.datetime.now()

        cash_post_id = latest_cash_post_id


        for unique_recon in unique_recon_list:

            matching = {}

            sub_invoice = 0

            sub_payment = 0

            sub_invoice_arr = []

            sub_payment_arr = []

            sub_check_arr = []

            multiple = []

            child_invoice = []

            child_payment = []

            child_ori_payment = []


            for invoice in imported_invoice_arr:            

                if unique_recon in invoice['recon_key']:

                    sub_invoice += float(invoice['invoice_amount'])

                    sub_invoice_arr.append(invoice['invoice'])

                    child_invoice.append(invoice)

                    multiple = invoice['recon_key']

            for recon in multiple:

                for payment in imported_payment_arr:

                    if recon in payment['recon_key']:

                        sub_payment += float(payment['check_amount'])

                        sub_payment_arr.append(str(payment['check_amount']) if '.' in str(payment['check_amount']) and len(str(payment['check_amount']).split('.')[1]) < 4 else  str(payment['check_amount']).split('.')[0] + '.' + str(payment['check_amount']).split('.')[1][:4] )

                        sub_check_arr.append(payment['payment_id'])

                        child_payment.append(payment)

                for payment in input_payment_arr:

                    if recon in payment['recon_key'] and payment not in child_ori_payment:

                        child_ori_payment.append(payment)

            # num = str(int(cash_post_id[2:])+1)

            # if len(num) < 4:

            #     for ind in range(0, 4-len(num)):

            #         num = '0'+num

            # cash_post_id = cash_post_id[:2]+num

            comment = []

            collection_status = ''

            current_cash_post = Cash_Post.objects.filter(recon_key=', '.join(multiple))

            if len(current_cash_post) == 0 :

                # comment = ''

                collection_status = ''

            else :

                # comment = current_cash_post[0].comment

                collection_status = current_cash_post[0].collection_status

            comment_list = Comment.objects.filter(recon_key=', '.join(multiple)).order_by('-posted_timestamp')

            for rec in comment_list:

                temp = model_to_dict(rec)

                logo_temp = temp['posted_by'].split(' ')[0][0].upper()

                try:
                    logo_temp += temp['posted_by'].split(' ')[1][0].upper()

                except:

                    pass

                temp['logo'] = logo_temp

                comment.append(temp) 

            a_r_status = ''


            balance_arr = []

            balance_db = Outstanding_Balance.objects.filter(cash_post_id=cash_post_id)

            for balance_row in balance_db:

                balance_arr.append(model_to_dict(balance_row))


            previous_balance = 0

            try:

                previous_balance_db = Outstanding_Balance.objects.filter(cash_post_id=cash_post_id).latest('posted_date')

                if previous_balance_db['posted_date'] == '{0.month}/{0.day}/{0.year}'.format(now):

                    previous_balance = float(previous_balance_db['amount'])

            except:

                previous_balance = sub_invoice

            difference = float(previous_balance - sub_payment)

            if difference == 0 : 

                a_r_status = 'paid in full'

            elif difference > 0 and difference <= 0.5:

                a_r_status = 'small balance'

            elif difference == sub_payment :

                a_r_status = 'unpaid'

            else :

                a_r_status = 'partial paid'

            
            matching = {

                'cash_post_id' : cash_post_id,

                'recon_key' : ', '.join(multiple),

                'invoice' : ', '.join(sub_invoice_arr),

                'payment' : ', '.join(sub_payment_arr),

                'payment_id' : ', '.join(sub_check_arr),

                'invoice_amount' : str(sub_invoice),

                'payment_amount' : str(sub_payment),

                'difference' : str(difference),

                'posted_date' : '{0.month}/{0.day}/{0.year}'.format(now),

                'child_invoice' : child_invoice,

                'child_payment' : child_payment,

                'child_ori_payment' : child_ori_payment,

                'comment' : comment,

                'a_r_status' : a_r_status,

                'collection_status' : collection_status,

                'balance_arr' : balance_arr

            }

            if difference > 0:


                check = Cash_Post.objects.filter(recon_key = ', '.join(multiple), posted_date__icontains='{0.month}/{0.day}/{0.year}'.format(now))

                if len(check) > 0:

                    matching['cash_post_id'] = check[0].cash_post_id


                for imported_payment in imported_payment_arr:

                    if unique_recon in imported_payment['recon_key']:

                        imported_payment['cash_post_id_memo'] = cash_post_id +'^'+imported_payment['recon_key']

                        check = Payment.objects.filter(payment_id=imported_payment['payment_id'])

                        if len(check) == 0:

                            payment_model = Payment(**imported_payment)

                            payment_model.save()

                        else :

                            temp = model_to_dict(check[0])

                            temp['cash_post_id_memo'] = imported_payment['cash_post_id_memo']
                        
                            check.update(**temp)


                request.session['imported_payment_arr'] = json.dumps(imported_payment_arr)

                matching_by_recon_key_arr.append(matching)

                data_db = {

                    'cash_post_id' : cash_post_id,

                    'recon_key' : ', '.join(multiple),

                    'invoice' : ', '.join(sub_invoice_arr),

                    'payment' : ', '.join(sub_payment_arr),

                    'payment_id' : ', '.join(sub_check_arr),

                    'invoice_amount' : str(sub_invoice),

                    'payment_amount' : str(sub_payment),

                    'difference' : str(difference),

                    'posted_date' : '{0.month}/{0.day}/{0.year}'.format(now),

                    'a_r_status' : a_r_status,

                    'collection_status' : collection_status

                }

                check = Cash_Post.objects.filter(recon_key = data_db['recon_key'])

                if len(check) == 0:

                    cash_post_model = Cash_Post(**data_db)

                    cash_post_model.save()

                else :

                    check.update(**data_db)


                balance_data = {

                    'cash_post_id' : cash_post_id,

                    'posted_date' : '{0.month}/{0.day}/{0.year}'.format(now),

                    'amount' : difference
                }



                check = Outstanding_Balance.objects.filter(cash_post_id = cash_post_id, posted_date=balance_data['posted_date'])

                if len(check) == 0:

                    balance_model = Outstanding_Balance(**balance_data)

                    balance_model.save()

                else :

                    check.update(**balance_data)


                num_temp = str(int(cash_post_id[2:])+1)

                if len(num_temp) < 4:

                    for ind in range(0, 4-len(num_temp)):

                        num_temp = '0'+num_temp

                cash_post_id = cash_post_id[:2] + num_temp

    except:

        pass

    request.session['matching_by_recon_key_arr'] = json.dumps(matching_by_recon_key_arr)

    return render(request, 'report/matching_by_recon_key.html',
        {

            'matching_by_recon_key_arr' : json.loads(request.session.get('matching_by_recon_key_arr', '[]')),

            'cash_post_batch_no_arr' : request.session.get('cash_post_batch_no_arr', '[]'),

            'selected_cash_post_id' : request.session.get('selected_cash_post_id', ''),

            'user' : request.session.get('user')

        })


def matching_by_recon_key(request):

    default_module(request)
    
    selected_cash_posts = []

    selected_cash_posts = Cash_Post.objects.all()

    matching_by_recon_key_arr = []

    for matching in selected_cash_posts:

        matching_by_recon_key_arr.append(model_to_dict(matching))
       
    new_matching_arr = []

    imported_payment_arr = json.loads(request.session.get('imported_payment_arr', '[]'))

    imported_invoice_arr = json.loads(request.session.get('imported_invoice_arr', '[]'))

    input_payment_arr = json.loads(request.session.get('input_payment_arr', '[]'))

    for matching in matching_by_recon_key_arr:

        child_invoice = []

        child_payment = []

        child_ori_payment = []

        multiple = []

        for invoice in imported_invoice_arr:            

            if matching['recon_key'] in invoice['recon_key']:

                child_invoice.append(invoice)

                multiple = invoice['recon_key']

        for recon in multiple:

            for payment in imported_payment_arr:

                if recon in payment['recon_key']:

                    child_payment.append(payment)

            for payment in input_payment_arr:

                if recon in payment['recon_key'] and payment not in child_ori_payment:

                    child_ori_payment.append(payment)

        comment = []

        comment_list = Comment.objects.filter(recon_key=matching['recon_key']).order_by('-posted_timestamp')

        for rec in comment_list:

            temp = model_to_dict(rec)

            logo_temp = temp['posted_by'].split(' ')[0][0].upper()

            try:
                logo_temp += temp['posted_by'].split(' ')[1][0].upper()

            except:

                pass

            temp['logo'] = logo_temp

            comment.append(temp)


        matching['child_invoice'] = child_invoice

        matching['child_payment'] = child_payment

        matching['child_ori_payment'] = child_ori_payment


        balance_arr = []

        balance_db = Outstanding_Balance.objects.filter(cash_post_id=matching['cash_post_id'])

        for balance_row in balance_db:

            balance_arr.append(model_to_dict(balance_row))

        matching['balance_arr'] = balance_arr

        matching['comment'] = comment

        new_matching_arr.append(matching)


    request.session['matching_by_recon_key_arr'] = json.dumps(new_matching_arr)


    return render(request, 'report/matching_by_recon_key.html',
        {

            'matching_by_recon_key_arr' : json.loads(request.session.get('matching_by_recon_key_arr', '[]')),

            'cash_post_batch_no_arr' : request.session.get('cash_post_batch_no_arr', '[]'),

            'user' : request.session.get('user')

        })


def report(request):

    default_module(request)

    return render(request, 'report/index.html',

        {

            'total_invoice_amount' : request.session.get('total_invoice_amount', ''),

            'reconned_invoice_amount' : request.session.get('reconned_invoice_amount', ''),

            'unused_invoice_amount' : request.session.get('unused_invoice_amount', ''),

            'total_payment_amount' : request.session.get('total_payment_amount', ''),

            'reconned_payment_amount' : request.session.get('reconned_payment_amount', ''),

            'unused_payment_amount' : request.session.get('unused_payment_amount', ''),

            'balance_on_recon_invoice' : request.session.get('balance_on_recon_invoice', '')

        })


def signup(request):
    
    if request.POST:

        uploaded_file_url = ''

        try:

            myfile = request.FILES['myfile']

            fs = FileSystemStorage()

            filename = fs.save('uploaded_files/images/'+myfile.name, myfile)

            uploaded_file_url = fs.url(filename)

        except:

            pass

        account = Account()

        account.firstname = request.POST["firstname"]

        account.lastname = request.POST["lastname"]

        account.username = request.POST["username"]

        account.email = request.POST["email"]

        account.password = request.POST["password"]

        account.image = uploaded_file_url

        account.role = "user"

        request.session['user'] = {
            # "id": user[0].id,
            "firstname" : request.POST["firstname"],

            "lastname" : request.POST["lastname"],

            "username": request.POST["username"],

            "email": request.POST["email"],

            "password": request.POST["password"],

            "image" : uploaded_file_url,

	        "role" : "user",
	    }

        account.save()
	   
        return render(request, 'login.html')

    return render(request, 'signup.html')


def changePwd(request):
    if request.POST:
        if request.POST["opassword"] == request.session['user']['password']:
            if request.POST["password"] == request.POST["rpassword"]:
                account = Account.objects.filter(email=request.session['user']['email']).update(password=request.POST["password"])
                request.session['user']['password'] = request.POST['password']
                return redirect("/home") 


    return render(request, 'changePwd.html', locals())


def forgot(request):
    if request.POST:
        if request.POST["password"] == request.POST["rpassword"]:
            account = Account.objects.filter(email=request.POST["email"]).update(password=request.POST["password"])
            return redirect("login") 

    return render(request, 'forgot.html')


def parse_superior_template(request, items):

    serial_number = 'XXXXXXXXXXX'

    parsed_payment_arr = []

    primary_cols = ['payment_id', 'check', 'recon_key', 'check_amount', 'payment_type', 'patient_name', 

        'visit_date', 'claim_id', 'payment_date', 'explanation_code', 'external_file_name', 'extra'

        ]

    input_files = items
    
    main_table_row = 2

    print('-------- Failed PDF files ---------\n')  

    for input_file_name in input_files:

        input_file_name = 'uploaded_files/'+input_file_name

        pdfFileObj = open( input_file_name, 'rb' )


        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

        num = pdfReader.numPages

        data = ''

        template = { }

        start_page = pdfReader.getPage(0).extractText()

        if 'superior' in start_page.lower():

            various_type = 'edge1'

            edge1 = 'ServDateProc#ModifiersDays/Ct/QtyCharged/AllowedDeductCoPayCoinsurDiscount/InterestMedAllow/MedPaidThirdPartyPayerDeniedEXPLCodesPayment/Withheld'

            edge2 = "ServDateProc#ModifiersDays/Ct/QtyChargedAllowedDeduct/CoPayDisallow/Add'tlPayInterest/DiscountMedAllow/MedPaidThirdPartyPayerDeniedEXPLCodesPayment/Withheld"

            if edge1 in pdfReader.getPage(0).extractText():

                various_type = 'edge1'

            if edge2 in pdfReader.getPage(0).extractText():

                various_type = 'edge2'

            data = 'InsuredName'+'InsuredName'.join(pdfReader.getPage(0).extractText().split('InsuredName')[1:]).replace(edge1, '').replace(edge2, '')

            for ind in range(1, num):

                edge = ''

                page = pdfReader.getPage(ind).extractText()

                if 'ServDate' in page[page.find('PaymentAmt'):page.find('PaymentAmt') + 50] : 

                    edge = page.split('ServDate')[0]

                if 'InsuredName' in page[page.find('PaymentAmt'):page.find('PaymentAmt') + 50] : 

                    edge = page.split('InsuredName')[0]

                data += pdfReader.getPage(ind).extractText().replace(edge, '').replace(edge1, '').replace(edge2,'')


            paragraph_list = data.split('InsuredName')

            for paragraph in paragraph_list[1:]:
                template['insured_name'] = getvalue(paragraph, ':', 'MbrNo')
                template['nbr_no'] = getvalue(paragraph, 'MbrNo:', 'MRN:')
                template['mrn'] = getvalue(paragraph, 'MRN:', 'Claim/CtrlNo:')
                template['clain_ctrl_no'] = getvalue(paragraph, 'Claim/CtrlNo:', 'PatientName:')
                template['patient_name'] = getvalue(paragraph, 'PatientName:', 'SvcProvNo:')
                template['svc_prov_no'] = getvalue(paragraph, 'SvcProvNo:', 'Carrier:')
                template['carrier'] = getvalue(paragraph, 'Carrier:', 'PatCtrlNo:')
                template['pat_ctrl_no'] = getvalue(paragraph, 'PatCtrlNo:', 'ServicingProvider:')
                template['servicing_provider'] = getvalue(paragraph, 'ServicingProvider:', 'NPI:')
                template['npi'] = getvalue(paragraph, 'NPI:', 'Group:')
                template['group'] = getvalue(paragraph, 'Group:', '010')

                paragraph = paragraph.replace('.000B', '.00 0B')

                record_list = paragraph.split('.000')

                if len(record_list) > 1:
                    for r_ind in range(0, len(record_list)):
                        try:

                            template['payment_id'] = 'EOB' + str( serial_number )

                            items = []
                            if r_ind == 0:
                                items = record_list[r_ind].split(template['group'])[1].split('$')
                                try:
                                    template['serv']=items[0].split('  ')[0][:4]
                                    template['date']=items[0].split('  ')[0][4:-7].split('/')[0] + '/' + items[0].split('  ')[0][4:-7].split('/')[1] + '/' + items[0].split('  ')[0][4:-7].split('/')[2][:4]
                                except:
                                    template['serv']=items[0].split(' ')[0][:4]
                                    template['date']=items[0].split('  ')[0][4:-9].split('/')[0] + '/' + items[0].split('  ')[0][4:-9].split('/')[1] + '/' + items[0].split('  ')[0][4:-9].split('/')[2][:4]
                                
                            else :
                                items = record_list[r_ind].split('$')
                                try:
                                    template['serv']='0'+items[0].split('  ')[0][:3]
                                    template['date']=items[0].split('  ')[0][3:-7].split('/')[0] + '/' + items[0].split('  ')[0][3:-7].split('/')[1] + '/' + items[0].split('  ')[0][3:-7].split('/')[2][:4]
                                except:
                                    template['serv']='0'+items[0].split(' ')[0][:3]
                                    template['date']=items[0].split('  ')[0][4:-8].split('/')[0] + '/' + items[0].split('  ')[0][4:-8].split('/')[1] + '/' + items[0].split('  ')[0][4:-8].split('/')[2][:4]

                            try:
                                template['proc #']=items[0].split('  ')[0][-7:-2]
                                template['modifiers']=items[0].split('  ')[0][-2:]
                                template['days_ct_qty']=items[0].split('  ')[1]
                            except:
                                template['proc #']=items[0].split(' ')[0][-9:-4]
                                template['modifiers']=items[0].split(' ')[0][-4:]
                                template['days_ct_qty']=items[0].split(' ')[1]


                            if various_type == 'edge1':
                                template['charged']='$'+items[1]
                                template['allowed']='$'+items[2]
                                template['deduct']='$'+items[3]
                                template['copay']='$'+items[4]
                                template['coinsur']='$'+items[5]
                                template['discount']='$'+items[6]
                                template['interest']='$'+items[7]
                                template['med_allow']='$'+items[8]
                                template['med_paid']='$'+items[9]
                                template['third_party_payer']='$'+items[10]
                                template['denied']='$'+items[11].strip()[:-2]
                                template['expl_codes']=items[11].strip()[-2:]
                                if 'JA' in template['denied']:
                                    template['denied'] = template['denied'].replace('JA', '')
                                    template['expl_codes'] = 'JA'+template['expl_codes']
                                template['payment']='$'+items[12]
                                template['withheld']='$'+items[13][:4]

                            if various_type == 'edge2' :
                                template['charged']='$'+items[1]
                                template['allowed']='$'+items[2]
                                template['deduct']='$'+items[3]
                                template['copay']='$'+items[4]
                                template['disallow']='$'+items[5]
                                template['addtl_pay']='$'+items[6]
                                template['interest']='$'+items[7]
                                template['discount']='$'+items[8]
                                template['med_allow']='$'+items[9]
                                template['med_paid']='$'+items[10]
                                template['third_party_payer']='$'+items[11]
                                template['denied']='$'+items[12].strip()[:-2]
                                template['expl_codes']=items[12].strip()[-2:]
                                if 'JA' in template['denied']:
                                    template['denied'] = template['denied'].replace('JA', '')
                                    template['expl_codes'] = 'JA'+template['expl_codes']
                                template['payment']='$'+items[13]
                                template['withheld']='$'+items[14][:4]


                            template['file_name'] = input_file_name.split('/')[1] 

                            data = {

                                'batch_no' : '',

                                'uploaded_date' : '',

                                'payment_id' : '',
                            
                                'cash_post_id_memo' : '',

                                'check' : '',

                                'recon_key' : template['pat_ctrl_no'],

                                'check_amount' : template['payment'],

                                'payment_type' : '',

                                'patient_name' : template['patient_name'],

                                'visit_date' : template['date'],

                                'claim_id' : template['clain_ctrl_no'],

                                'payment_date' : '',

                                'explanation_code' : template['expl_codes'],

                                'external_file_name' : input_file_name,

                                'extra' : template

                            }

                            parsed_payment_arr.append(data)

                            # for key, value in template.items():

                            #     if '$' in  value and '.pdf' not in value.lower():

                            #         try:

                            #             value = float(value.replace('$', ''))

                            #         except :

                            #             value = 0

                            #     main_table_sheet.cell( column = main_table_headers.index(key) + 1, row = main_table_row , value = value).number_format = '$#,##0.00'

                        except :

                            pass

                else:

                    template['payment_id'] = 'EOB' + str( serial_number )

                    items = paragraph.split(template['group'])[1].split('Sub-total')[0].split('$')

                    try:
                        template['serv']=items[0].split('  ')[0][:4]
                        template['date']=items[0].split('  ')[0][4:-7]
                        template['proc #']=items[0].split('  ')[0][-7:-2]
                        template['modifiers']=items[0].split('  ')[0][-2:]
                        template['days_ct_qty']=items[0].split('  ')[1]

                    except:

                        template['serv']=items[0].split(' ')[0][:4]
                        template['date']=items[0].split(' ')[0][4:-9]
                        template['proc #']=items[0].split(' ')[0][-9:-4]
                        template['modifiers']=items[0].split(' ')[0][-4:]
                        template['days_ct_qty']=items[0].split(' ')[1]

                    if various_type == 'edge1':
                        template['charged']='$'+items[1]
                        template['allowed']='$'+items[2]
                        template['deduct']='$'+items[3]
                        template['copay']='$'+items[4]
                        template['coinsur']='$'+items[5]
                        template['discount']='$'+items[6]
                        template['interest']='$'+items[7]
                        template['med_allow']='$'+items[8]
                        template['med_paid']='$'+items[9]
                        template['third_party_payer']='$'+items[10]
                        template['denied']='$'+items[11].strip()[:-2]
                        template['expl_codes']=items[11].strip()[-2:]
                        if 'JA' in template['denied']:
                            template['denied'] = template['denied'].replace('JA', '')
                            template['expl_codes'] = 'JA'+template['expl_codes']
                        template['payment']='$'+items[12]
                        template['withheld']='$'+items[13][:4]

                    if various_type == 'edge2' :
                        template['charged']='$'+items[1]
                        template['allowed']='$'+items[2]
                        template['deduct']='$'+items[3]
                        template['copay']='$'+items[4]
                        template['disallow']='$'+items[5]
                        template['addtl_pay']='$'+items[6]
                        template['interest']='$'+items[7]
                        template['discount']='$'+items[8]
                        template['med_allow']='$'+items[9]
                        template['med_paid']='$'+items[10]
                        template['third_party_payer']='$'+items[11]
                        template['denied']='$'+items[12].strip()[:-2]
                        template['expl_codes']=items[12].strip()[-2:]
                        if 'JA' in template['denied']:
                            template['denied'] = template['denied'].replace('JA', '')
                            template['expl_codes'] = 'JA'+template['expl_codes']
                        template['payment']='$'+items[13]
                        template['withheld']='$'+items[14][:4]


                    template['file_name'] = input_file_name.split('/')[1]

                    data = {

                        'batch_no' : 'XXXXXXXXXXXXXXXXXXXx',

                        'uploaded_date' : 'VVVVVVVVVVVVVVV',

                        'payment_id' : 'CCCCCCCCCC',
                    
                        'cash_post_id_memo' : 'bBBBBBBBBBBB',

                        'check' : '',

                        'recon_key' : template['pat_ctrl_no'],

                        'check_amount' : template['payment'],

                        'payment_type' : '',

                        'patient_name' : template['patient_name'],

                        'visit_date' : template['date'],

                        'claim_id' : template['clain_ctrl_no'],

                        'payment_date' : '',

                        'explanation_code' : template['expl_codes'],

                        'external_file_name' : input_file_name,

                        'extra' : json.dumps(template)

                    }


                    parsed_payment_arr.append(data)

                    # check = Payment_All.objects.filter(check=data['check'], recon_key=data['recon_key'], check_amount=data['check_amount'])

                    # if len(check) == 0:

                    payment_model = Payment_All(**data)

                    payment_model.save()

                    # else :

                    #     data['batch_no'] = check[0].batch_no

                    #     data['cash_post_id_memo'] = check[0].cash_post_id_memo

                    #     check.update(**data)

                    # for key, value in template.items():

                    #     if '$' in  value and '.pdf' not in value.lower():

                    #         try:

                    #             value = float(value.replace('$', ''))

                    #         except :

                    #             value = 0



                        # main_table_sheet.cell( column = main_table_headers.index(key) + 1, row = main_table_row , value = value).number_format = '$#,##0.00'

        else :

            print( input_file_name.split('/')[1])


        return parsed_payment_arr
